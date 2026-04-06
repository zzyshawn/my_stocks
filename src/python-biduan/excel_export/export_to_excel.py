"""
将缠论分析结果导出为Excel文档
"""

import pandas as pd
from datetime import datetime
import os
from typing import Dict, Any, List
from chanlun.core.kline import KLine
from chanlun.core.analyzer import ChanLunAnalyzer
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


import time


def export_analysis_to_excel(analysis_result: Dict[str, Any], filename: str = "analysis_result.xlsx", max_retries: int = 3, retry_delay: int = 10):
    """
    将缠论分析结果导出为Excel文档
    
    Args:
        analysis_result: 缠论分析结果
        filename: 输出文件名
        max_retries: 最大重试次数
        retry_delay: 重试延迟（秒）
    """
    # 获取数据
    kline_series = analysis_result['kline_series']
    fractals = analysis_result['fractals']
    kline_states = analysis_result.get('kline_states', [])
    
    # 从核心代码获取有效分型数据（1.3.3强制要求：直接使用核心代码的计算结果）
    valid_top_fractals = analysis_result.get('valid_top_fractals', set())
    valid_bottom_fractals = analysis_result.get('valid_bottom_fractals', set())
    top_bottom_markers = analysis_result.get('top_bottom_markers', {})
    
    # 创建数据框架
    data = {
        '日期': [],
        '开盘价': [],
        '最高价': [],
        '最低价': [],
        '收盘价': [],
        '成交量': [],
        'K线状态': [],  # 上升、下降、上升包含、下降包含
        '合并K线状态': [],  # 合并后的K线状态，只标记为上升、下降
        '包含起始日期': [],  # 包含关系的起始日期
        '分型类型': [],  # 顶分型、底分型、无
        '有效分型': [],  # 顶顶分型、底底分型、无
        '有效分型组合': [],  # 有效顶底组合、有效底顶组合、有效底底组合、有效顶顶组合、无
        '顶底': [],  # 上顶、下底、无
        '前顶底日期': [],  # 前一个顶底标记的日期
        '中枢顶ZG': [],  # 笔中枢顶
        '中枢底ZD': [],  # 笔中枢底
        'ZG时间': [],  # 中枢顶时间
        'ZD时间': []   # 中枢底时间
    }
    
    # 获取分型索引
    top_fractals = set(fractals['top'])
    bottom_fractals = set(fractals['bottom'])
    
    # 获取包含关系映射
    inclusion_mapping = analysis_result.get('inclusion_mapping', {})
    
    # 创建包含关系起始日期映射
    inclusion_start_dates = {}
    for processed_idx, original_indices in inclusion_mapping.items():
        if len(original_indices) > 1:  # 只有当多个原始K线合并成一个处理后K线时才有包含关系
            # 包含关系的起始日期是第一个原始K线的日期
            start_idx = min(original_indices)
            start_timestamp = kline_series._sorted_klines[start_idx].timestamp
            # 判断是否包含时间信息
            if start_timestamp.hour == 0 and start_timestamp.minute == 0 and start_timestamp.second == 0:
                # 日线数据，只显示日期
                start_date = start_timestamp.strftime('%Y-%m-%d')
            else:
                # 分钟数据，显示日期和时间
                start_date = start_timestamp.strftime('%Y-%m-%d %H:%M:%S')
            for idx in original_indices:
                inclusion_start_dates[idx] = start_date
    
    # 为每个原始K线确定合并后的状态
    merged_kline_states = []
    
    # 遍历所有原始K线
    for i, kline in enumerate(kline_series._sorted_klines):
        # 找到包含该原始K线的处理后K线
        merged_state = ""
        for processed_idx, original_indices in inclusion_mapping.items():
            if i in original_indices:
                # 找到处理后K线对应的状态
                # 由于处理后K线的状态是根据包含关系处理后的结果
                # 我们需要根据原始K线的状态来推断合并后的状态
                # 对于包含关系中的K线，合并后的状态应该是包含关系处理后的方向
                # 这里我们简单地使用第一个原始K线的状态（非包含）作为合并后的状态
                for orig_idx in original_indices:
                    if orig_idx < len(kline_states):
                        orig_state = kline_states[orig_idx]
                        if orig_state == "上升" or orig_state == "下降":
                            merged_state = orig_state
                            break
                if not merged_state:
                    # 如果没有找到非包含状态，则使用当前K线的状态
                    if i < len(kline_states):
                        current_state = kline_states[i]
                        if current_state == "上升" or current_state == "下降":
                            merged_state = current_state
                        elif current_state == "上升包含":
                            merged_state = "上升"
                        elif current_state == "下降包含":
                            merged_state = "下降"
                break
        
        merged_kline_states.append(merged_state)
    
    # 创建顶分型和底分型的有序列表，按索引排序
    sorted_top_fractals = sorted(list(top_fractals))
    sorted_bottom_fractals = sorted(list(bottom_fractals))
    
    # 识别有效分型组合（基于核心代码返回的valid_top_fractals和valid_bottom_fractals）
    # 创建所有有效分型的有序列表，按索引排序
    all_valid_fractals = sorted(list(valid_top_fractals | valid_bottom_fractals))
    
    # 识别有效分型组合：有效顶底组合、有效底顶组合、有效底底组合、有效顶顶组合
    valid_combinations = {}  # key: 索引, value: (组合类型, 组合定义)
    
    for i, idx in enumerate(all_valid_fractals):
        if i > 0:  # 从第二个有效分型开始
            prev_idx = all_valid_fractals[i-1]  # 前一个有效分型的索引
            
            # 判断前一个有效分型的类型
            prev_type = None
            if prev_idx in valid_top_fractals:
                prev_type = "顶顶分型"
            elif prev_idx in valid_bottom_fractals:
                prev_type = "底底分型"
            
            # 判断当前有效分型的类型
            curr_type = None
            if idx in valid_top_fractals:
                curr_type = "顶顶分型"
            elif idx in valid_bottom_fractals:
                curr_type = "底底分型"
            
            # 识别有效分型组合
            combination_type = None
            combination_definition = None
            if prev_type == "顶顶分型" and curr_type == "底底分型":
                combination_type = "有效顶底组合"
                combination_definition = "有效顶底组合：前一个有效分型是顶顶分型，当前是底底分型"
            elif prev_type == "底底分型" and curr_type == "顶顶分型":
                combination_type = "有效底顶组合"
                combination_definition = "有效底顶组合：前一个有效分型是底底分型，当前是顶顶分型"
            elif prev_type == "底底分型" and curr_type == "底底分型":
                combination_type = "有效底底组合"
                combination_definition = "有效底底组合：前一个有效分型是底底分型，当前是底底分型"
            elif prev_type == "顶顶分型" and curr_type == "顶顶分型":
                combination_type = "有效顶顶组合"
                combination_definition = "有效顶顶组合：前一个有效分型是顶顶分型，当前是顶顶分型"
            
            if combination_type:
                valid_combinations[idx] = (combination_type, combination_definition)
    
    # 1.3.3强制要求：直接使用核心代码返回的top_bottom_markers，不重新计算
    
    # 1.3.3强制要求：直接使用核心代码返回的prev_top_bottom_dates，不重新计算
    prev_top_bottom_dates = analysis_result.get('prev_top_bottom_dates', {})
    
    # 1.3.3强制要求：直接使用核心代码返回的笔中枢数据，不重新计算
    zg_dict = analysis_result.get('zg_dict', {})
    zd_dict = analysis_result.get('zd_dict', {})
    zg_time_dict = analysis_result.get('zg_time_dict', {})
    zd_time_dict = analysis_result.get('zd_time_dict', {})
    
    # 遍历所有K线，填充数据
    for i, kline in enumerate(kline_series._sorted_klines):
        # 确定K线状态
        kline_state = ""
        if i < len(kline_states):
            kline_state = kline_states[i]
        
        # 确定合并K线状态
        merged_kline_state = ""
        if i < len(merged_kline_states):
            merged_kline_state = merged_kline_states[i]
        
        # 确定包含起始日期
        inclusion_start_date = ""
        if i in inclusion_start_dates:
            inclusion_start_date = inclusion_start_dates[i]
        
        # 确定分型类型
        fractal_type = ""
        if i in top_fractals:
            fractal_type = "顶分型"
        elif i in bottom_fractals:
            fractal_type = "底分型"
        else:
            fractal_type = "无"
        
        # 确定有效分型
        valid_fractal_type = ""
        if i in valid_top_fractals:
            valid_fractal_type = "顶顶分型"
        elif i in valid_bottom_fractals:
            valid_fractal_type = "底底分型"
        else:
            valid_fractal_type = "无"
        
        # 确定有效分型组合
        valid_combination_type = ""
        if i in valid_combinations:
            valid_combination_type, _ = valid_combinations[i]  # 只使用组合类型，不使用定义
        else:
            valid_combination_type = "无"
        
        # 确定顶底标记
        top_bottom_marker = top_bottom_markers.get(i, "无")
        
        # 确定前顶底日期
        prev_top_bottom_date = prev_top_bottom_dates.get(i, "")
        
        # 添加数据
        # 判断是否包含时间信息
        if kline.timestamp.hour == 0 and kline.timestamp.minute == 0 and kline.timestamp.second == 0:
            # 日线数据，只显示日期
            date_str = kline.timestamp.strftime('%Y-%m-%d')
        else:
            # 分钟数据，显示日期和时间
            date_str = kline.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        data['日期'].append(date_str)
        data['开盘价'].append(kline.open)
        data['最高价'].append(kline.high)
        data['最低价'].append(kline.low)
        data['收盘价'].append(kline.close)
        data['成交量'].append(kline.volume)
        data['K线状态'].append(kline_state)
        data['合并K线状态'].append(merged_kline_state)
        data['包含起始日期'].append(inclusion_start_date)
        data['分型类型'].append(fractal_type)
        data['有效分型'].append(valid_fractal_type)
        data['有效分型组合'].append(valid_combination_type)
        data['顶底'].append(top_bottom_marker)
        data['前顶底日期'].append(prev_top_bottom_date)
        
        # 确定笔中枢数据
        zg = zg_dict.get(i)
        zd = zd_dict.get(i)
        zg_time = zg_time_dict.get(i)
        zd_time = zd_time_dict.get(i)
        
        data['中枢顶ZG'].append(zg if zg is not None else "")
        data['中枢底ZD'].append(zd if zd is not None else "")
        
        # 格式化ZG时间
        if zg_time:
            if zg_time.hour == 0 and zg_time.minute == 0 and zg_time.second == 0:
                zg_time_str = zg_time.strftime('%Y-%m-%d')
            else:
                zg_time_str = zg_time.strftime('%Y-%m-%d %H:%M:%S')
        else:
            zg_time_str = ""
        data['ZG时间'].append(zg_time_str)
        
        # 格式化ZD时间
        if zd_time:
            if zd_time.hour == 0 and zd_time.minute == 0 and zd_time.second == 0:
                zd_time_str = zd_time.strftime('%Y-%m-%d')
            else:
                zd_time_str = zd_time.strftime('%Y-%m-%d %H:%M:%S')
        else:
            zd_time_str = ""
        data['ZD时间'].append(zd_time_str)
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 保存到Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='缠论分析结果', index=False)
        
        # 获取工作表对象，以便设置格式
        worksheet = writer.sheets['缠论分析结果']
        
        # 定义背景填充
        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # 上升包含 - 黄色
        green_fill = PatternFill(start_color="0000FF00", end_color="0000FF00", fill_type="solid")    # 下降包含 - 绿色
        
        # 为有包含关系的行设置背景色
        # 获取包含关系信息，以便同时标记包含和被包含的K线
        inclusion_info = analysis_result.get('inclusion_info', [])
        
        # 创建包含关系映射，找出需要标记颜色的行
        inclusion_rows = set()
        
        # 从包含关系信息中标记相关行
        for idx, state in inclusion_info:
            if "包含" in state:
                inclusion_rows.add(idx + 2)  # Excel行号从1开始，加上标题行，所以要+2
        
        # 标记包含关系中的所有行（包含方和被包含方）
        inclusion_mapping = analysis_result.get('inclusion_mapping', {})
        for processed_idx, original_indices in inclusion_mapping.items():
            if len(original_indices) > 1:  # 说明有多根原始K线合并成一根处理后K线
                for original_idx in original_indices:
                    excel_row = original_idx + 2  # Excel行号从1开始，加上标题行，所以要+2
                    if len(original_indices) > 1:  # 这些都是包含关系中的K线
                        # 检查这些K线的状态来确定颜色
                        if original_idx < len(kline_states) and "包含" in kline_states[original_idx]:
                            inclusion_rows.add(excel_row)
        
        # 为有包含关系的行设置背景色
        # 获取包含关系信息，以便同时标记包含和被包含的K线
        inclusion_info = analysis_result.get('inclusion_info', [])
        
        # 创建包含关系映射，找出需要标记颜色的行
        inclusion_rows = set()
        
        # 从包含关系信息中标记相关行
        for idx, state in inclusion_info:
            if "包含" in state:
                inclusion_rows.add(idx + 2)  # Excel行号从1开始，加上标题行，所以要+2
        
        # 标记包含关系中的所有行（包含方和被包含方）
        inclusion_mapping = analysis_result.get('inclusion_mapping', {})
        for processed_idx, original_indices in inclusion_mapping.items():
            if len(original_indices) > 1:  # 说明有多根原始K线合并成一根处理后K线
                # 这些都是包含关系中的K线，都需要标记颜色
                for original_idx in original_indices:
                    excel_row = original_idx + 2  # Excel行号从1开始，加上标题行，所以要+2
                    inclusion_rows.add(excel_row)
        
        # 为包含关系的行设置背景色
        for row_idx, kline_state in enumerate(kline_states):
            # Excel的行号从1开始，加上标题行，所以要+2
            excel_row = row_idx + 2
            
            if kline_state == "上升包含":
                # 为上升包含设置黄色背景
                for col in range(1, len(df.columns) + 1):  # 列号从1开始
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.fill = yellow_fill
            elif kline_state == "下降包含":
                # 为下降包含设置绿色背景
                for col in range(1, len(df.columns) + 1):  # 列号从1开始
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.fill = green_fill
        
        # 根据包含关系映射额外标记被包含的K线行
        for excel_row in inclusion_rows:
            # 检查这一行的K线状态来确定颜色
            row_idx = excel_row - 2  # 转换回数组索引
            if row_idx < len(kline_states):
                kline_state = kline_states[row_idx]
                if "包含" in kline_state:
                    # 已经标记过的跳过
                    continue
                else:
                    # 这是被包含的K线，需要根据包含它的K线状态来标记颜色
                    # 由于这些行在包含关系中，我们根据包含它们的K线来确定颜色
                    # 但目前我们无法确定具体颜色，所以我们需要通过映射来确定
                    for processed_idx, original_indices in inclusion_mapping.items():
                        if row_idx in original_indices and len(original_indices) > 1:
                            # 找到包含关系，确定颜色
                            # 检查这些K线中是否有被标记为包含的，以确定颜色
                            for orig_idx in original_indices:
                                if orig_idx < len(kline_states) and "包含" in kline_states[orig_idx]:
                                    # 找到了标记为包含的K线，使用相同颜色
                                    color_kline_state = kline_states[orig_idx]
                                    for col in range(1, len(df.columns) + 1):
                                        cell = worksheet.cell(row=excel_row, column=col)
                                        if "上升包含" in color_kline_state:
                                            cell.fill = yellow_fill
                                        elif "下降包含" in color_kline_state:
                                            cell.fill = green_fill
                                    break
                            break
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"分析结果已导出到: {filename}")


def _process_kline_inclusion_with_labels(kline_series):
    """
    处理K线包含关系并标记K线类型
    
    Args:
        kline_series: K线序列
        
    Returns:
        (标记了类型的K线列表, 包含关系索引列表)
    """
    if len(kline_series) <= 1:
        labels = ["向上" if k.close >= k.open else "向下" for k in kline_series._sorted_klines]
        return labels, []
    
    labels = ["向上" if k.close >= k.open else "向下" for k in kline_series._sorted_klines]
    inclusion_indices = set()  # 存储有包含关系的索引
    
    # 标记包含关系
    i = 1
    while i < len(kline_series):
        current_kline = kline_series[i-1]
        next_kline = kline_series[i]
        
        # 检查是否存在包含关系
        if (current_kline.high >= next_kline.high and current_kline.low <= next_kline.low):  # 左包含
            # 当前K线包含下一个K线，根据当前K线的方向确定类型
            current_direction = 1 if current_kline.close >= current_kline.open else -1
            if current_direction == 1:  # 向上K线
                labels[i-1] = "向上"
                labels[i] = "向上包含"
            else:  # 向下K线
                labels[i-1] = "向下"
                labels[i] = "向下包含"
            
            # 记录包含关系的索引
            inclusion_indices.add(i-1)
            inclusion_indices.add(i)
        elif (next_kline.high >= current_kline.high and next_kline.low <= current_kline.low):  # 右包含
            # 下一个K线包含当前K线，根据当前K线的方向确定类型
            current_direction = 1 if current_kline.close >= current_kline.open else -1
            if current_direction == 1:  # 向上K线
                labels[i-1] = "向上"
                labels[i] = "向上包含"
            else:  # 向下K线
                labels[i-1] = "向下"
                labels[i] = "向下包含"
            
            # 记录包含关系的索引
            inclusion_indices.add(i-1)
            inclusion_indices.add(i)
        
        i += 1
    
    return labels, sorted(list(inclusion_indices))


def analyze_and_export_single_stock(stock_code: str, klines: List[KLine], output_dir: str = None, data_type: str = None):
    """
    分析单个股票并导出Excel
    
    Args:
        stock_code: 股票代码
        klines: K线数据
        output_dir: 输出目录，默认为股票数据的对应目录
        data_type: 数据类型（如 'd', '30min', '5min'），用于区分输出文件名
    """
    # 如果没有指定输出目录，则使用股票数据的对应目录
    if output_dir is None:
        output_dir = f"X:\\股票信息\\股票数据库\\daban\\股票数据\\{stock_code}"
    
    # 确保目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    # 创建分析器
    analyzer = ChanLunAnalyzer()
    
    # 执行分析
    result = analyzer.analyze(klines)
    
    # 根据数据类型生成文件名
    if data_type:
        filename = os.path.join(output_dir, f"{stock_code}_{data_type}_analysis.xlsx")
    else:
        filename = os.path.join(output_dir, f"{stock_code}_analysis.xlsx")
    
    # 导出到Excel
    export_analysis_to_excel(result, filename)
    
    return result


if __name__ == "__main__":
    print("Excel导出功能模块已准备就绪")