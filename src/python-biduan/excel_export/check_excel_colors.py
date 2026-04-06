"""
检查Excel文件中的颜色标记情况
"""

import os
from openpyxl import load_workbook


def check_excel_colors():
    """
    检查Excel文件中的颜色标记情况
    """
    stock_code = "603906"
    data_type = "d"
    
    # 导出的Excel文件路径
    excel_file = f"X:\\股票信息\\股票数据库\\daban\\股票数据\\{stock_code}\\{stock_code}_{data_type}_analysis.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Excel文件不存在: {excel_file}")
        return
    
    print(f"检查Excel文件: {excel_file}")
    
    # 加载Excel文件
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # 检查标题行
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"Excel表头: {headers}")
    
    # 查找K线状态列
    kline_state_col = None
    for i, header in enumerate(headers):
        if header == "K线状态":
            kline_state_col = i + 1  # Excel列号从1开始
            break
    
    if not kline_state_col:
        print("未找到K线状态列")
        return
    
    print(f"K线状态列位置: {kline_state_col}")
    
    # 统计不同状态的行数和颜色标记情况
    state_count = {}
    color_marked_count = {}
    
    # 遍历数据行
    for row in range(2, ws.max_row + 1):  # 从第二行开始（第一行是表头）
        # 获取K线状态
        state_cell = ws.cell(row=row, column=kline_state_col)
        state = state_cell.value
        
        if state not in state_count:
            state_count[state] = 0
            color_marked_count[state] = 0
        
        state_count[state] += 1
        
        # 检查是否有背景色
        has_color = False
        cell = ws.cell(row=row, column=1)
        if cell.fill.start_color.rgb and cell.fill.start_color.rgb != "00000000":
            has_color = True
            color_marked_count[state] += 1
    
    # 输出统计结果
    print("\nK线状态统计:")
    for state, count in state_count.items():
        marked_count = color_marked_count.get(state, 0)
        print(f"{state}: {count} 行, 其中 {marked_count} 行有颜色标记")
    
    # 检查一些具体的包含关系示例
    print("\n检查包含关系示例:")
    for row in range(2, ws.max_row + 1):
        state_cell = ws.cell(row=row, column=kline_state_col)
        state = state_cell.value
        
        if "包含" in state:
            # 检查这一行的颜色
            cell = ws.cell(row=row, column=1)
            color_rgb = cell.fill.start_color.rgb
            
            # 检查前一行和后一行是否也被标记
            prev_row = row - 1
            next_row = row + 1
            
            print(f"行 {row-1} (Excel行 {row}): 状态={state}, 颜色={color_rgb}")
            
            if prev_row >= 2:
                prev_state = ws.cell(row=prev_row, column=kline_state_col).value
                prev_color = ws.cell(row=prev_row, column=1).fill.start_color.rgb
                print(f"  前一行: 状态={prev_state}, 颜色={prev_color}")
            
            if next_row <= ws.max_row:
                next_state = ws.cell(row=next_row, column=kline_state_col).value
                next_color = ws.cell(row=next_row, column=1).fill.start_color.rgb
                print(f"  后一行: 状态={next_state}, 颜色={next_color}")
            
            print()
            
            # 只显示前5个示例
            if len([s for s in state_count.keys() if "包含" in s]) > 0:
                if state_count.get("上升包含", 0) + state_count.get("下降包含", 0) > 5:
                    if state_count.get(state, 0) > 5:
                        break
    
    print("颜色检查完成")


if __name__ == "__main__":
    check_excel_colors()
